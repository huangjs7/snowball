"""生成图 4：六个典型时刻的整条板凳龙位置分布。"""

from __future__ import annotations

import argparse
from pathlib import Path
import re
import sys

import matplotlib.pyplot as plt
from matplotlib.collections import LineCollection
from matplotlib.lines import Line2D
import numpy as np
from openpyxl import load_workbook


HERE = Path(__file__).resolve().parent
FIGURES_ROOT = HERE.parent
PROJECT_ROOT = HERE.parents[2]
sys.path.insert(0, str(FIGURES_ROOT))

import common_plotting as plotting  # noqa: E402


TIMES = (0, 60, 120, 180, 240, 300)
NODE_COUNT = 224
HEAD_BENCH_LENGTH = 3.41
BODY_BENCH_LENGTH = 2.20
HEAD_HANDLE_DISTANCE = 2.86
BODY_HANDLE_DISTANCE = 1.65
OUTPUT_STEM = "图04_典型时刻整龙位置分布"


def locate_result_file(explicit_path: Path | None) -> Path:
    """优先使用命令行路径，其次查找仓库内的填充结果文件。"""

    candidates = []
    if explicit_path is not None:
        candidates.append(explicit_path)
    candidates.extend(
        [
            PROJECT_ROOT / "提交" / "result1.xlsx",
            PROJECT_ROOT / "建模计算" / "问题1" / "输出" / "result1.xlsx",
        ]
    )
    for path in candidates:
        resolved = path.expanduser().resolve()
        if resolved.is_file():
            return resolved
    checked = "\n".join(f"- {path}" for path in candidates)
    raise FileNotFoundError(f"未找到填充后的 result1.xlsx，已检查：\n{checked}")


def parse_time_columns(ws) -> dict[int, int]:
    columns: dict[int, int] = {}
    for column in range(2, ws.max_column + 1):
        header = ws.cell(row=1, column=column).value
        match = re.fullmatch(r"\s*(-?\d+)\s*s\s*", str(header))
        if match:
            columns[int(match.group(1))] = column
    return columns


def load_snapshots(path: Path) -> dict[int, np.ndarray]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    if "位置" not in workbook.sheetnames:
        raise ValueError("result1.xlsx 缺少“位置”工作表")
    ws = workbook["位置"]
    columns = parse_time_columns(ws)
    missing = [time for time in TIMES if time not in columns]
    if missing:
        raise ValueError(f"位置表缺少典型时刻列：{missing}")

    snapshots: dict[int, np.ndarray] = {}
    for time in TIMES:
        column = columns[time]
        points = np.empty((NODE_COUNT, 2), dtype=float)
        for index in range(NODE_COUNT):
            x_value = ws.cell(row=2 + 2 * index, column=column).value
            y_value = ws.cell(row=3 + 2 * index, column=column).value
            if x_value is None or y_value is None:
                raise ValueError(f"t={time} s 时第 {index} 个把手坐标为空")
            points[index] = float(x_value), float(y_value)
        if not np.isfinite(points).all():
            raise ValueError(f"t={time} s 的位置数据含非有限值")
        snapshots[time] = points
    workbook.close()
    return snapshots


def validate_chord_constraints(snapshots: dict[int, np.ndarray]) -> float:
    expected = np.array(
        [HEAD_HANDLE_DISTANCE] + [BODY_HANDLE_DISTANCE] * (NODE_COUNT - 2),
        dtype=float,
    )
    max_error = 0.0
    for points in snapshots.values():
        distances = np.linalg.norm(np.diff(points, axis=0), axis=1)
        max_error = max(max_error, float(np.max(np.abs(distances - expected))))
    if max_error > 1e-5:
        raise ValueError(f"相邻把手定长约束最大误差 {max_error:.3e} m，超过容差")
    return max_error


def spiral_arc_primitive(theta: np.ndarray) -> np.ndarray:
    """阿基米德螺线从极点到参数 theta 的弧长原函数。"""

    return 0.5 * plotting.B * (
        theta * np.sqrt(1.0 + theta**2) + np.arcsinh(theta)
    )


def validate_initial_condition(snapshots: dict[int, np.ndarray]) -> tuple[float, float]:
    """核验 t=0 时龙头前把手位于第 16 圈 A 点。"""

    actual = snapshots[0][0]
    expected = np.array([plotting.R_INITIAL, 0.0])
    position_error = float(np.linalg.norm(actual - expected))
    theta = float(np.linalg.norm(actual) / plotting.B)
    theta_error = abs(theta - plotting.THETA_INITIAL)
    if position_error > 1e-6 or theta_error > 1e-6:
        raise ValueError(
            "初始条件错误：t=0 龙头前把手必须位于 "
            "A(theta=32*pi, r=8.8 m)"
        )
    return position_error, theta_error


def validate_length_relations(
    snapshots: dict[int, np.ndarray],
) -> tuple[float, float, float]:
    """核验板凳实长 L、螺线弧长 s 与把手直线弦长 d 的次序。"""

    bench_lengths = np.array(
        [HEAD_BENCH_LENGTH] + [BODY_BENCH_LENGTH] * (NODE_COUNT - 2),
        dtype=float,
    )
    min_theta_increment = np.inf
    min_arc_minus_chord = np.inf
    min_bench_minus_arc = np.inf

    for time, points in snapshots.items():
        theta = np.linalg.norm(points, axis=1) / plotting.B
        theta_increment = np.diff(theta)
        if np.any(theta_increment <= 0):
            raise ValueError(f"t={time} s 时把手未从龙头向龙尾沿螺线向外排列")

        chord_lengths = np.linalg.norm(np.diff(points, axis=0), axis=1)
        arc_lengths = np.diff(spiral_arc_primitive(theta))
        min_theta_increment = min(
            min_theta_increment, float(np.min(theta_increment))
        )
        min_arc_minus_chord = min(
            min_arc_minus_chord, float(np.min(arc_lengths - chord_lengths))
        )
        min_bench_minus_arc = min(
            min_bench_minus_arc, float(np.min(bench_lengths - arc_lengths))
        )

    if min_arc_minus_chord <= 0 or min_bench_minus_arc <= 0:
        raise ValueError(
            "长度关系错误：应满足板凳实长 L > 螺线弧长 s > 直线弦长 d"
        )
    return min_theta_increment, min_arc_minus_chord, min_bench_minus_arc


def build_bench_segments(points: np.ndarray) -> np.ndarray:
    """按 3.41 m/2.20 m 实长生成每节板凳的直线中心段。"""

    chords = np.diff(points, axis=0)
    chord_lengths = np.linalg.norm(chords, axis=1)
    directions = chords / chord_lengths[:, None]
    midpoints = 0.5 * (points[:-1] + points[1:])
    bench_lengths = np.array(
        [HEAD_BENCH_LENGTH] + [BODY_BENCH_LENGTH] * (NODE_COUNT - 2),
        dtype=float,
    )
    half_vectors = 0.5 * bench_lengths[:, None] * directions
    return np.stack([midpoints - half_vectors, midpoints + half_vectors], axis=1)


def draw_figure(snapshots: dict[int, np.ndarray]) -> None:
    # 按 A4 正文约 16 cm 宽设计，避免论文缩放后字号过小。
    fig, axes = plt.subplots(2, 3, figsize=(8.0, 5.3), sharex=True, sharey=True)
    panel_tags = "abcdef"
    time_box_colors = [
        plotting.LIGHT_BLUE,
        plotting.LIGHT_CYAN,
        plotting.LIGHT_GREEN,
        plotting.LIGHT_PURPLE,
        plotting.LIGHT_ORANGE,
        plotting.LIGHT_RED,
    ]
    axis_limit = 12.45
    ticks = np.arange(-10, 11, 5)
    spiral_theta = np.linspace(0.02, axis_limit / plotting.B, 14000)
    spiral_x, spiral_y = plotting.spiral_xy(spiral_theta)

    for panel_index, (ax, time, box_color) in enumerate(
        zip(axes.flat, TIMES, time_box_colors)
    ):
        points = snapshots[time]

        ax.axhline(0, color=plotting.GRID, lw=0.8, zorder=0)
        ax.axvline(0, color=plotting.GRID, lw=0.8, zorder=0)
        ax.grid(color=plotting.GRID, lw=0.45, alpha=0.38, zorder=0)

        # 完整参考螺线必须可见：t=0 时 A 点内侧尚未被板凳龙占据。
        ax.plot(
            spiral_x,
            spiral_y,
            color=plotting.LIGHT_BLUE,
            lw=0.62,
            alpha=0.52,
            zorder=1,
        )

        # 板凳是两把手连线方向上的直线实体，并按 3.41 m/2.20 m 实长延伸。
        benches = LineCollection(
            build_bench_segments(points),
            colors=plotting.DARK_BLUE,
            linewidths=1.9,
            alpha=0.88,
            capstyle="butt",
            zorder=3,
        )
        ax.add_collection(benches)

        # 所有把手中心均用深蓝色；龙头、龙尾只增加彩色外圈作识别。
        ax.scatter(
            points[:, 0],
            points[:, 1],
            s=7.0,
            color=plotting.DARK_BLUE,
            edgecolors="white",
            linewidths=0.20,
            zorder=4,
        )
        ax.scatter(
            points[0, 0],
            points[0, 1],
            s=42,
            facecolor="none",
            edgecolor=plotting.DARK_RED,
            linewidth=1.25,
            zorder=6,
        )
        ax.scatter(
            points[-1, 0],
            points[-1, 1],
            s=39,
            facecolor="none",
            edgecolor=plotting.DARK_PURPLE,
            linewidth=1.25,
            zorder=6,
        )
        ax.scatter(0, 0, s=13, color=plotting.INK, zorder=5)
        ax.text(0.23, 0.22, "$O$", color=plotting.INK, ha="left", va="bottom", zorder=7)

        if time == 0:
            ax.annotate(
                "$A$（第 16 圈）\n$\\theta=32\\pi,\\ r=8.8\\,\\mathrm{m}$",
                xy=points[0],
                xytext=(4.25, -2.65),
                arrowprops=dict(arrowstyle="->", color=plotting.DARK_RED, lw=0.9),
                bbox=dict(
                    boxstyle="round,pad=0.25",
                    facecolor=plotting.LIGHT_YELLOW,
                    edgecolor="none",
                    alpha=0.90,
                ),
                color=plotting.DARK_RED,
                fontsize=7.6,
                ha="center",
                va="top",
                zorder=9,
            )

        ax.text(
            0.035,
            0.955,
            rf"({panel_tags[panel_index]})  $t={time}\,\mathrm{{s}}$",
            transform=ax.transAxes,
            ha="left",
            va="top",
            color=plotting.INK,
            weight="bold",
            bbox=dict(
                boxstyle="round,pad=0.24",
                facecolor=box_color,
                edgecolor="none",
                alpha=0.56,
            ),
            zorder=8,
        )

        ax.set_xlim(-axis_limit, axis_limit)
        ax.set_ylim(-axis_limit, axis_limit)
        ax.set_aspect("equal", adjustable="box")
        ax.set_xticks(ticks)
        ax.set_yticks(ticks)
        ax.tick_params(length=2.5, width=0.6, color="#777777")
        ax.spines["top"].set_visible(False)
        ax.spines["right"].set_visible(False)
        ax.spines["left"].set_color("#b7b7b7")
        ax.spines["bottom"].set_color("#b7b7b7")
        ax.spines["left"].set_linewidth(0.65)
        ax.spines["bottom"].set_linewidth(0.65)

    legend_items = [
        Line2D([0], [0], color=plotting.LIGHT_BLUE, lw=1.3, label="等距螺线"),
        Line2D(
            [0],
            [0],
            color=plotting.DARK_BLUE,
            lw=1.9,
            marker="o",
            markerfacecolor=plotting.DARK_BLUE,
            markeredgecolor="white",
            markersize=4.2,
            label="直线板凳与把手中心",
        ),
        Line2D(
            [0],
            [0],
            marker="o",
            color="none",
            markerfacecolor=plotting.DARK_BLUE,
            markeredgecolor=plotting.DARK_RED,
            markeredgewidth=1.2,
            markersize=6.3,
            label="龙头前把手",
        ),
        Line2D(
            [0],
            [0],
            marker="o",
            color="none",
            markerfacecolor=plotting.DARK_BLUE,
            markeredgecolor=plotting.DARK_PURPLE,
            markeredgewidth=1.2,
            markersize=6.0,
            label="龙尾后把手",
        ),
    ]
    fig.supxlabel(r"$x\,/\,\mathrm{m}$", y=0.062)
    fig.supylabel(r"$y\,/\,\mathrm{m}$", x=0.035)
    fig.legend(
        handles=legend_items,
        loc="lower center",
        ncol=4,
        frameon=False,
        bbox_to_anchor=(0.5, 0.012),
        handlelength=1.7,
        columnspacing=1.25,
    )
    fig.subplots_adjust(left=0.075, right=0.992, top=0.985, bottom=0.135, wspace=0.10, hspace=0.16)
    plotting.save_figure(fig, OUTPUT_STEM)


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", type=Path, help="填充后的 result1.xlsx 路径")
    args = parser.parse_args()

    plotting.OUTPUT_DIR = HERE
    plotting.configure_style()
    source = locate_result_file(args.input)
    snapshots = load_snapshots(source)
    max_error = validate_chord_constraints(snapshots)
    position_error, theta_error = validate_initial_condition(snapshots)
    theta_step, arc_gap, bench_gap = validate_length_relations(snapshots)
    draw_figure(snapshots)
    print(f"图 4 已生成：{HERE}")
    print(f"数据源：{source}")
    print(f"节点数：{NODE_COUNT}；典型时刻：{TIMES}")
    print(f"定长约束最大误差：{max_error:.3e} m")
    print(
        "初始 A 点误差："
        f"位置 {position_error:.3e} m；极角 {theta_error:.3e} rad"
    )
    print(f"龙头至龙尾的最小极角增量：{theta_step:.3e} rad")
    print(
        "长度关系 L>s>d："
        f"min(s-d)={arc_gap:.3e} m；min(L-s)={bench_gap:.3e} m"
    )


if __name__ == "__main__":
    main()
